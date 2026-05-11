import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), 'lib'))

import boto3
import json
import traceback
from datetime import datetime, timezone
from io import BytesIO
import openpyxl
from collections import defaultdict

def lambda_handler(event, context):
    """
    Lambda handler for Security Hub Excel report generation.
    
    Args:
        event: Lambda event
        context: Lambda context
        
    Returns:
        Dictionary with execution result
    """
    try:
        print(f"Starting Security Hub report generation at {datetime.now().isoformat()}")
        
        # Get environment variables
        bucket_name = os.environ.get('S3_BUCKET_NAME')
        
        # For debugging
        print(f"Environment variables:")
        print(f"S3_BUCKET_NAME: {bucket_name}")
        
        # Override from event if provided
        if event and isinstance(event, dict):
            bucket_name = event.get('bucket_name', bucket_name)
        
        if not bucket_name:
            return {
                'statusCode': 400,
                'body': json.dumps({
                    'message': 'S3_BUCKET_NAME environment variable is required',
                    'error': 'Missing configuration'
                })
            }
            
        print(f"Using bucket: {bucket_name}")
        
        # Generate today's filename
        today = datetime.now()
        timestamp = today.strftime("%Y%m%d_%H%M%S")
        output_filename = f"security_hub_report_{timestamp}.xlsx"
        
        # Create a memory file object for the Excel report
        memory_file = BytesIO()
        
        # Get AWS session for the Lambda execution environment
        session = boto3.Session()
        
        # Initialize Security Hub client
        print("Initializing Security Hub client")
        securityhub = session.client("securityhub")
        
        # Check if Security Hub is enabled
        try:
            print("Checking Security Hub status")
            hub_status = securityhub.describe_hub()
            print(f"Security Hub ARN: {hub_status.get('HubArn', 'Unknown')}")
        except Exception as e:
            print(f"Error checking Security Hub status: {str(e)}")
            return {
                'statusCode': 400,
                'body': json.dumps({
                    'message': 'Security Hub is not enabled or accessible',
                    'error': str(e)
                })
            }
        
        # Get Security Hub findings
        print("Retrieving Security Hub findings...")
        findings = get_security_hub_findings(securityhub)
        print(f"Retrieved {len(findings)} findings")
        
        if not findings:
            print("No findings retrieved from Security Hub")
            
        # Transform findings into structured data
        print("Transforming findings data...")
        transformed_findings = transform_findings_data(findings)
        
        # Create summary statistics
        print("Creating summary statistics...")
        summary_stats = create_summary_statistics(transformed_findings)
        
        # Create workbook
        print("Creating Excel workbook...")
        wb = openpyxl.Workbook()
        
        # Create Executive Summary sheet
        create_summary_sheet(wb, summary_stats, len(findings))
        
        # Create Detailed Findings sheet
        create_findings_sheet(wb, transformed_findings)
        
        # Create Pivot Tables sheet
        create_pivot_sheet(wb, transformed_findings)
        
        # Save to memory file
        print("Saving workbook to memory")
        wb.save(memory_file)
        memory_file.seek(0)
        
        # Upload to S3
        print(f"Uploading to S3 bucket: {bucket_name}")
        s3_client = session.client('s3')
        s3_key = f"reports/{output_filename}"
        
        s3_client.upload_fileobj(
            memory_file,
            bucket_name,
            s3_key,
            ExtraArgs={
                'ContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Metadata': {
                    'generated-by': 'security-hub-excel-generator',
                    'generated-at': datetime.now(timezone.utc).isoformat(),
                    'content-type': 'security-hub-report'
                }
            }
        )
        
        print(f"Report uploaded to s3://{bucket_name}/{s3_key}")
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Security Hub Excel report generated successfully',
                'bucket': bucket_name,
                'key': s3_key,
                'findings_count': len(findings),
                'timestamp': timestamp,
                'worksheets_created': ['Executive Summary', 'Detailed Findings', 'Pivot Analysis']
            })
        }
        
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        traceback.print_exc()
        return {
            'statusCode': 500,
            'body': json.dumps({
                'message': f'Error generating Security Hub report: {str(e)}',
                'error': traceback.format_exc()
            })
        }

def get_security_hub_findings(securityhub):
    """
    Retrieve findings from AWS Security Hub
    
    Args:
        securityhub: boto3 Security Hub client
        
    Returns:
        List of findings
    """
    findings = []
    paginator = securityhub.get_paginator('get_findings')
    
    # Filter for active findings from any security standard
    filters = {
        'RecordState': [
            {
                'Value': 'ACTIVE',
                'Comparison': 'EQUALS'
            }
        ]
    }
    
    try:
        print("Starting to retrieve Security Hub findings...")
        
        for page_num, page in enumerate(paginator.paginate(
            Filters=filters,
            PaginationConfig={'MaxItems': 1000, 'PageSize': 100}
        ), 1):
            page_findings = page.get('Findings', [])
            findings.extend(page_findings)
            print(f"Retrieved {len(page_findings)} findings from page {page_num}")
            
            # Log progress for large datasets
            if len(findings) % 500 == 0:
                print(f"Total findings retrieved so far: {len(findings)}")
                
    except Exception as e:
        print(f"Error retrieving Security Hub findings: {str(e)}")
        raise
        
    print(f"Total findings retrieved: {len(findings)}")
    return findings

def transform_findings_data(findings):
    """
    Transform Security Hub findings into structured data
    
    Args:
        findings: List of Security Hub findings
        
    Returns:
        List of dictionaries with structured finding data
    """
    print("Transforming findings to structured data...")
    transformed_data = []
    
    for finding in findings:
        # Extract resource information (handle multiple resources)
        resources = finding.get('Resources', [])
        primary_resource = resources[0] if resources else {}
        
        # Extract compliance information
        compliance = finding.get('Compliance', {})
        
        # Extract severity information
        severity = finding.get('Severity', {})
        
        # Extract remediation information
        remediation = finding.get('Remediation', {}).get('Recommendation', {})
        
        # Build the row data
        row = {
            'Finding_ID': finding.get('Id', 'N/A'),
            'Title': finding.get('Title', 'N/A'),
            'Description': (finding.get('Description', 'N/A')[:500] + '...') if len(finding.get('Description', '')) > 500 else finding.get('Description', 'N/A'),
            'Severity_Label': severity.get('Label', 'N/A'),
            'Severity_Score': severity.get('Normalized', 0),
            'Compliance_Status': compliance.get('Status', 'N/A'),
            'Compliance_Status_Reasons': ', '.join([reason.get('ReasonCode', '') for reason in compliance.get('StatusReasons', [])]),
            'Resource_Type': primary_resource.get('Type', 'N/A'),
            'Resource_ID': primary_resource.get('Id', 'N/A'),
            'Resource_Region': primary_resource.get('Region', 'N/A'),
            'AWS_Account_ID': finding.get('AwsAccountId', 'N/A'),
            'Region': finding.get('Region', 'N/A'),
            'Created_At': finding.get('CreatedAt', 'N/A'),
            'Updated_At': finding.get('UpdatedAt', 'N/A'),
            'First_Observed_At': finding.get('FirstObservedAt', 'N/A'),
            'Last_Observed_At': finding.get('LastObservedAt', 'N/A'),
            'Generator_ID': finding.get('GeneratorId', 'N/A'),
            'Product_Name': finding.get('ProductFields', {}).get('aws/securityhub/ProductName', 'N/A'),
            'Company_Name': finding.get('ProductFields', {}).get('aws/securityhub/CompanyName', 'N/A'),
            'Remediation_URL': remediation.get('Url', 'N/A'),
            'Remediation_Text': (remediation.get('Text', 'N/A')[:200] + '...') if len(remediation.get('Text', '')) > 200 else remediation.get('Text', 'N/A'),
            'Workflow_Status': finding.get('Workflow', {}).get('Status', 'N/A'),
            'Record_State': finding.get('RecordState', 'N/A'),
            'Source_URL': finding.get('SourceUrl', 'N/A'),
            'Resource_Count': len(resources)
        }
        transformed_data.append(row)
    
    print(f"Transformed {len(transformed_data)} findings into structured data")
    return transformed_data

def create_summary_statistics(transformed_findings):
    """
    Create summary statistics from transformed findings
    
    Args:
        transformed_findings: List of transformed finding dictionaries
        
    Returns:
        Dictionary with summary statistics
    """
    print("Creating summary statistics...")
    summary_data = {}
    
    if not transformed_findings:
        return summary_data
    
    # Overall statistics
    summary_data['total_findings'] = len(transformed_findings)
    
    # Severity breakdown
    severity_counts = defaultdict(int)
    for finding in transformed_findings:
        severity_counts[finding['Severity_Label']] += 1
    summary_data['severity_breakdown'] = dict(severity_counts)
    
    # Compliance status breakdown
    compliance_counts = defaultdict(int)
    for finding in transformed_findings:
        compliance_counts[finding['Compliance_Status']] += 1
    summary_data['compliance_breakdown'] = dict(compliance_counts)
    
    # Workflow status breakdown
    workflow_counts = defaultdict(int)
    for finding in transformed_findings:
        workflow_counts[finding['Workflow_Status']] += 1
    summary_data['workflow_breakdown'] = dict(workflow_counts)
    
    # Top resource types
    resource_counts = defaultdict(int)
    for finding in transformed_findings:
        resource_counts[finding['Resource_Type']] += 1
    summary_data['top_resources'] = dict(sorted(resource_counts.items(), key=lambda x: x[1], reverse=True)[:10])
    
    # Account distribution
    account_counts = defaultdict(int)
    for finding in transformed_findings:
        account_counts[finding['AWS_Account_ID']] += 1
    summary_data['account_distribution'] = dict(account_counts)
    
    return summary_data

def create_summary_sheet(wb, summary_stats, total_findings):
    """
    Create the Executive Summary worksheet
    
    Args:
        wb: openpyxl workbook
        summary_stats: Dictionary with summary statistics
        total_findings: Total number of findings
    """
    # Remove default sheet and create Executive Summary
    if 'Sheet' in [ws.title for ws in wb.worksheets]:
        wb.remove(wb['Sheet'])
    
    ws = wb.create_sheet("Executive Summary")
    
    # Add header
    ws.append(["Security Hub Compliance Report"])
    ws.append([f"Generated: {datetime.now().isoformat()}"])
    ws.append([f"Total Findings: {total_findings}"])
    ws.append([""])
    
    if not summary_stats:
        ws.append(["No findings data available"])
        return
    
    # Severity breakdown
    ws.append(["Severity Breakdown"])
    ws.append(["Severity Level", "Count", "Percentage"])
    for severity, count in summary_stats.get('severity_breakdown', {}).items():
        percentage = f"{(count / total_findings * 100):.1f}%" if total_findings > 0 else "0.0%"
        ws.append([severity, count, percentage])
    ws.append([""])
    
    # Compliance status breakdown
    ws.append(["Compliance Status Breakdown"])
    ws.append(["Status", "Count", "Percentage"])
    for status, count in summary_stats.get('compliance_breakdown', {}).items():
        percentage = f"{(count / total_findings * 100):.1f}%" if total_findings > 0 else "0.0%"
        ws.append([status, count, percentage])
    ws.append([""])
    
    # Workflow status breakdown
    ws.append(["Workflow Status Breakdown"])
    ws.append(["Status", "Count", "Percentage"])
    for status, count in summary_stats.get('workflow_breakdown', {}).items():
        percentage = f"{(count / total_findings * 100):.1f}%" if total_findings > 0 else "0.0%"
        ws.append([status, count, percentage])
    ws.append([""])
    
    # Top resource types
    ws.append(["Top Resource Types"])
    ws.append(["Resource Type", "Count", "Percentage"])
    for resource_type, count in summary_stats.get('top_resources', {}).items():
        percentage = f"{(count / total_findings * 100):.1f}%" if total_findings > 0 else "0.0%"
        ws.append([resource_type, count, percentage])

def create_findings_sheet(wb, transformed_findings):
    """
    Create the Detailed Findings worksheet
    
    Args:
        wb: openpyxl workbook
        transformed_findings: List of transformed finding dictionaries
    """
    ws = wb.create_sheet("Detailed Findings")
    
    if not transformed_findings:
        ws.append(["No findings available"])
        return
    
    # Add headers (use keys from first finding)
    headers = list(transformed_findings[0].keys())
    ws.append(headers)
    
    # Add data rows
    for finding in transformed_findings:
        row = [finding.get(header, 'N/A') for header in headers]
        ws.append(row)
    
    print(f"Added {len(transformed_findings)} rows to Detailed Findings sheet")

def create_pivot_sheet(wb, transformed_findings):
    """
    Create the Pivot Analysis worksheet
    
    Args:
        wb: openpyxl workbook
        transformed_findings: List of transformed finding dictionaries
    """
    ws = wb.create_sheet("Pivot Analysis")
    
    if not transformed_findings:
        ws.append(["No findings available for analysis"])
        return
    
    # Severity by Resource Type analysis
    ws.append(["Severity by Resource Type"])
    
    # Get unique resource types and severities
    resource_types = set(f['Resource_Type'] for f in transformed_findings)
    severities = set(f['Severity_Label'] for f in transformed_findings)
    
    # Create header row
    header = ['Resource Type'] + sorted(list(severities))
    ws.append(header)
    
    # Create data rows
    for resource_type in sorted(resource_types):
        row = [resource_type]
        for severity in sorted(severities):
            count = sum(1 for f in transformed_findings 
                       if f['Resource_Type'] == resource_type and f['Severity_Label'] == severity)
            row.append(count)
        ws.append(row)
    
    ws.append([])
    ws.append(["Compliance Status by Resource Type"])
    
    # Get unique compliance statuses
    compliance_statuses = set(f['Compliance_Status'] for f in transformed_findings)
    
    # Create header row
    header = ['Resource Type'] + sorted(list(compliance_statuses))
    ws.append(header)
    
    # Create data rows
    for resource_type in sorted(resource_types):
        row = [resource_type]
        for status in sorted(compliance_statuses):
            count = sum(1 for f in transformed_findings 
                       if f['Resource_Type'] == resource_type and f['Compliance_Status'] == status)
            row.append(count)
        ws.append(row)

# For local testing
if __name__ == "__main__":
    # Set up environment variables for local testing
    os.environ['S3_BUCKET_NAME'] = 'your-test-bucket-name'
    
    # Mock event and context for local testing
    test_event = {}
    test_context = type('obj', (object,), {
        'function_name': 'test_function',
        'function_version': '1',
        'aws_request_id': 'test-request-id'
    })
    
    # Run the function
    result = lambda_handler(test_event, test_context)
    print(json.dumps(result, indent=2))